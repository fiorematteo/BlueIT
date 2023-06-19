import asyncio
import time
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from string import ascii_uppercase
from copy import deepcopy

font_name: str = "Inherit"
border_color: str = "5B9BD5"
light_fill = PatternFill(
		fill_type="solid",
		start_color='deeaf6',
		end_color='deeaf6'
		)
dark_fill = PatternFill(
		fill_type="solid",
		start_color='5b9bd5',
		end_color='5b9bd5'
		)
border = Border(
	top=Side(
		border_style="thick",
		color=border_color
		),
	right=Side(
		border_style="thick",
		color=border_color
		),
	bottom=Side(
		border_style="thick",
		color=border_color
		),
	left=Side(
		border_style="thick",
		color=border_color
		)
	)
alignment = Alignment(
	horizontal='center',
	vertical='center',
	wrap_text=True
	)


async def get_date() -> str:
	t = time.localtime(time.time())
	date_time = datetime.datetime(t.tm_year, t.tm_mon, t.tm_mday, 9, 0)
	today: int = int(time.mktime(date_time.timetuple()))
	week_unix: int = 604800
	last_week: int = today - week_unix
	lw = time.localtime(last_week)

	return f"{lw.tm_mday}/{lw.tm_mon}/{lw.tm_year} - {t.tm_mday}/{t.tm_mon}/{t.tm_year}"


async def wr_xlsx(offenses_list: list, file_path: str) -> None:
	wb: Workbook = Workbook()
	ws = wb.active
	ws.append(["Resoconto settimanale S.O.C."])
	r = ws["A1"]
	r.font = Font(
		name=font_name,
		size=10,
		bold=False,
		color='FF000000'
		)
	r.fill = light_fill
	r.border = border
	r.alignment = alignment

	ws.append([await get_date()])
	r = ws["A2"]
	r.font = Font(
		name=font_name,
		size=16,
		bold=True
		)
	r.fill = dark_fill
	r.border = border
	r.alignment = alignment

	ws.append([""])

	next_key: str = ""
	keys_list: list = list(offenses_list[0].keys())
	counters: list = [keys_list[0]]
	for idx, key in enumerate(offenses_list[0].keys()):
		if key == keys_list[0] or key == keys_list[-1] or key == next_key:
			continue
		next_key = keys_list[idx + 1]
		show_value: bool = offenses_list[0][next_key]
		if show_value:
			counters.append(key)
	counters.append(keys_list[-1])
	ws.append(counters)
	for idx in range(len(counters)):
		r = ws[f'{ascii_uppercase[idx]}4']
		r.font = Font(
			name=font_name,
			size=10,
			bold=bool(idx),
			color='FF000000'
			)
		r.fill = dark_fill
		r.border = border
		r.alignment = alignment
	values: list = []
	for key in counters:
		values.append(offenses_list[0][key])
	ws.append(values)
	for idx in range(len(counters)):
		r = ws[f'{ascii_uppercase[idx]}5']
		r.font = Font(
			name=font_name,
			size=10,
			bold=bool(idx),
			color='FF0000'
			)
		r.fill = light_fill
		r.border = border
		r.alignment = alignment

	ws.merge_cells("A6:D24")

	ws.append([""])
	mylist: list = ["ID", "Data/Ora", "Minaccia", "Categoria"]
	for idx in range(4):
		r = ws[f'{ascii_uppercase[idx]}25']
		r.value = mylist[idx]
		r.font = Font(
			name=font_name,
			size=10,
			bold=bool(idx),
			color='FF000000'
			)
		r.fill = dark_fill
		r.border = border
		r.alignment = alignment

	offenses_list.pop(0)
	row: int = 26
	for offense in offenses_list:
		ws.append([offense["id"], offense["start_time"], offense["description"]])
		for idx in range(4):
			r = ws[f'{ascii_uppercase[idx]}{row}']
			r.font = Font(
				name=font_name,
				size=7,
				bold=False,
				color='FF000000'
				)
			r.fill = light_fill
			r.border = border
			r.alignment = alignment
		row += 1
		ws.append([offense["note"]])
		ws.merge_cells(f"A{row}:D{row}")
		for idx in range(4):
			r = ws[f'{ascii_uppercase[idx]}{row}']
			r.font = Font(
				name=font_name,
				size=7,
				bold=False,
				color='FF000000'
				)
			r.border = border
			r.alignment = Alignment(
				horizontal='left',
				vertical='center',
				wrap_text=True
				)
		row += 1

	wb.save(file_path)


if __name__ == "__main__":
	file_path: str = "C:\\Download\TEST.xlsx"
	mylist: list = [
		{"Allarmi totali": 50, "Breach alert di Darktrace": 30, "show_dt": True, "Allarmi di Sophos": 0, "show_sophos": False, "Offensive di QRadar": 20},
		{"id": 69, "start_time": "06/12/2020", "description": "Desc Lorem ipsum", "note": "Note Lorem ipsum"},
		{"id": 420, "start_time": "25/12/2020", "description": "Desc Lorem ipsum", "note": "Note Lorem ipsum"}
		]
	asyncio.run(wr_xlsx(mylist, file_path))
